import requests
import json
import tableprint
import os
import time
import sys
import openpyxl as ws
from config import *


requests.packages.urllib3.disable_warnings()

def get_onlinescans():
    count = 0
    list_scan_info = []
    response = requests.get(define.host + "/api/v1/scans", headers=define.api_header, timeout=30, verify=False)
    results = json.loads(response.content)
    for result in results['scans']:
        scan_info = {'target': result['target']['address'],
                     'description': result['target']['description'],
                     'scan_id': result['scan_id'],
                     'report_id': result['current_session']['scan_session_id'],
                     'status': result['current_session']['status']}
        list_scan_info.append(scan_info)
    for i in range(len(list_scan_info)):
        if list_scan_info[i]['status'] == 'processing':
            count = count + 1     
    return count

def get_scans():
    creat_xlsx()
    c = 0
    while True:
        result = requests.get(define.host+"api/v1/scans?c="+str(c),headers=define.api_header,timeout=30,verify=False)
        results = json.loads(result.content)
        c = c + 100
        if results['scans'] == []:
            return print(define.RED+"[*]任务执行完毕 文件保存为:%s"%define.filename)
        for s in results["scans"]:
            get_vulnerabilities(s['scan_id'],s['current_session']['scan_session_id'],s['target']['address'])


def get_vulnerabilities(scan_id,scan_session_id,host):
    c = 0
    while True:
        result = requests.get(url=define.host+"api/v1/scans/"+scan_id+"/results/"+scan_session_id+"/vulnerabilities?c=%s"%str(c),headers=define.api_header,timeout=30,verify=False)
        results = json.loads(result.text)
        c = c + 100
        if results['vulnerabilities'] == []:
            return print(define.BLUE+"[-]当前扫描抓取结束 scan_id:%s"%scan_id)
        for s in results['vulnerabilities']:
            get_details(scan_id,scan_session_id,host,s['vuln_id'])


def get_details(scan_id,scan_session_id,host,vuln_id):
    vulnerabilities = {}
    result = requests.get(url=define.host+"api/v1/scans/"+scan_id+"/results/"+scan_session_id+"/vulnerabilities/"+vuln_id,headers=define.api_header,timeout=30,verify=False)
    results = json.loads(result.text)
    vulnerabilities['details'] = results['details']
    vulnerabilities['request'] = results['request']
    vulnerabilities['affects_url'] = results['affects_url']
    vulnerabilities['description'] = results['description']
    vulnerabilities['vt_name'] = results['vt_name']
    vulnerabilities['recommendation'] = results['recommendation']
    vulnerabilities['severity'] = results['severity']
    vulnerabilities['host'] = host
    vulnerabilities['affects_detail'] = results['affects_detail']
    write_xlsx(vulnerabilities)


def write_xlsx(vulnerabilities):
    print(define.GREEN+"[*]内容正在写入 vu_name:%s"%vulnerabilities['vt_name'])
    wb = ws.load_workbook(define.filename)
    sheet1 = wb['Sheet']
    num = sheet1.max_row
    sheet1.cell(row = num+1,column = 1,value = vulnerabilities['host'])
    sheet1.cell(row = num+1,column = 2,value = vulnerabilities['vt_name'])
    sheet1.cell(row = num+1,column = 3,value = vulnerabilities['severity'])
    sheet1.cell(row = num+1,column = 4,value = vulnerabilities['affects_detail'])
    sheet1.cell(row = num+1,column = 5,value = vulnerabilities['affects_url'])
    sheet1.cell(row = num+1,column = 6,value = vulnerabilities['request'])
    sheet1.cell(row = num+1,column = 7,value = vulnerabilities['recommendation'])
    sheet1.cell(row = num+1,column = 8,value = vulnerabilities['description'])
    sheet1.cell(row = num+1,column = 9,value = vulnerabilities['details'])
    wb.save(define.filename)


def creat_xlsx():
    if os.path.exists(define.filename) == False:
        s = 0
        wb = ws.Workbook()
        ws1 = wb.active
        if os.path.exists('out/') == False:
            os.mkdir('out')
        word=['风险目标','风险名称','风险等级(3-0由高危到infomation)','风险参数','风险地址','风险请求','整改意见','风险描述','风险详情']
        for i in word:
            s = s + 1
            ws1.cell(row =1,column = s,value = i)
        wb.save(define.filename)
        print(define.RED+"[*]创建文件成功 %s"%define.filename)
    else:
        print(define.RED+"[*]文件已存在 文件为:%s"%define.filename)


x = []

def task(files):
    s = open('%s'%files,'r')
    for i in s.readlines():
        i = i.strip()
        x.append(i)
    s.close


def add(url):
    #添加任务
    data = {"address":url,"description":url,"criticality":"10"}
    try:
        response = requests.post(define.host+"api/v1/targets",data=json.dumps(data),headers=define.api_header,timeout=30,verify=False)
        result = json.loads(response.content)
        return result['target_id']
    except Exception as e:
        print(str(e))
        return


def single_scan(url,rule='full'):
    data = {'target_id':add(url),'profile_id':define.awvs_scan_rule[rule],
        'schedule':{'disable':False,'start_date':None, 'time_sensitive':False}}
    try:
        r = requests.post(url=define.host + 'api/v1/scans', timeout=10, verify=False, headers=define.api_header, data=json.dumps(data))
        if r.status_code == 201:
            print(define.BLUE + '[-] OK, 扫描任务已经启动 当前扫描:%s...'%url)
    except Exception as e:
        print (e)

def delete_all():
    c = 0
    print(define.RED+"[*]开始清除任务")
    while True:
        result = requests.get(define.host+"api/v1/targets?c="+str(c),headers=define.api_header,timeout=30,verify=False)
        results = json.loads(result.content)
        c = c + 100
        if results['targets'] == []:
            return print(define.RED+"[*]任务全部清除完毕")
        for s in results["targets"]:
            r = requests.delete(url=define.host+'api/v1/targets/'+s['target_id'], timeout=10, verify=False, headers=define.api_header)
            print(define.BLUE+"[-]当前删除 target_id:%s"%s['target_id'])

if __name__ == '__main__':
    print(define.ORANGE+define.banner)
    if len(sys.argv) < 2:
        print(define.ORANGE+define.usage)
    elif sys.argv[1] == '-f':
        try:
            if len(sys.argv) > 3:
                rule = sys.argv[3]
                task(str(sys.argv[2]))
                print(define.RED+"[*]扫描开始添加")
                for s in x:
                    while True:
                        if get_onlinescans() < define.maxscans:
                            single_scan(s, rule)
                            time.sleep(5)
                            break
                        print('[-] 当前已达到最大并发扫描数，等待5分钟后自动再次尝试！')
                        time.sleep(300)
                print(define.RED+"[*]扫描添加完毕")
            if len(sys.argv) == 3:
                task(str(sys.argv[2]))
                print(define.RED+"[*]扫描开始添加")
                for s in x:
                    while True:
                        if get_onlinescans() < define.maxscans:
                            single_scan(s)
                            time.sleep(5)
                            break
                        print('[-] 当前已达到最大并发扫描数，等待5分钟后自动再次尝试！')
                        time.sleep(300)
                print(define.RED+"[*]扫描添加完毕")
        except:
            print(define.BLUE+'    [*]Usage example: Python3 Acunetix12-Scan-Agent.py -f url.txt')
    elif sys.argv[1] == '-d':
        delete_all()
    elif sys.argv[1] == '-o':
        get_scans()

    else:
        print(define.ORANGE+define.usage)
